from __future__ import annotations

import io
import os
import threading
import time
import zipfile
from email.message import EmailMessage

import db
import persistence
import pytest
import photo_store


def _isolate_store(monkeypatch, tmp_path) -> None:
    # Queue worker metadata yolu persistence.STORE_PATH'ten dinamik çözülür.
    # Önceki testin worker thread'i tamamen kapanmadan yeni tmp_path'e geçmek,
    # üretimde mümkün olmayan testler-arası bir yol yarışı oluşturur.
    try:
        from backend import services

        deadline = time.monotonic() + 3
        while services._import_worker_alive and time.monotonic() < deadline:
            time.sleep(0.01)
    except (ImportError, AttributeError):
        pass
    monkeypatch.setattr(persistence, "STORE_PATH", str(tmp_path / "state.json"))
    monkeypatch.setattr(photo_store, "PHOTO_DIR", str(tmp_path / "photos"))
    monkeypatch.setattr(db, "_engine", None)
    monkeypatch.setattr(db, "_init_failed", True)


def _csv(passport: str, departure: str, name: str = "JOHN") -> bytes:
    return (
        "NO,NAME,SURNAME,PASSPORT NUMBER,VOUCHER,DEPARTURE,ARRIVAL,ADULT,CHILD\n"
        f"1,{name},DOE,{passport},V1,{departure},2026-07-22,25,0\n"
    ).encode("utf-8")


def _xlsx(passport: str, departure: str, name: str = "JOHN") -> bytes:
    """Üretimdeki gerçek Gate Visa yerleşimiyle küçük bir XLSX üretir."""
    from openpyxl import load_workbook
    from gate_visa_reader import build_gate_visa_template_xlsx

    workbook = load_workbook(io.BytesIO(build_gate_visa_template_xlsx()))
    sheet = workbook.active
    values = [1, name, "DOE", passport, "V-1", departure, "2026-12-31", 25, 0]
    for column, value in enumerate(values, start=1):
        sheet.cell(row=5, column=column, value=value)
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def test_composite_dedup_date_scope_package_and_undo(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    from backend import services

    first = services.import_gate_visa_files([("first.csv", _csv("U12345678", "2026-07-15"))], batch_id="batch-1", dup_strategy="skip")
    another_day = services.import_gate_visa_files([("second.csv", _csv("U12345678", "2026-07-16"))], batch_id="batch-1", dup_strategy="skip")
    duplicate = services.import_gate_visa_files([("duplicate.csv", _csv("U12345678", "2026-07-15"))], batch_id="batch-1", dup_strategy="skip")

    assert first[0] == 1
    assert another_day[0] == 1
    assert duplicate[0] == 0
    assert duplicate[5] == 1
    assert services.get_summary().passenger_count == 2
    assert services.get_summary("Aralık", "2026-07-15", "2026-07-15").passenger_count == 1

    with pytest.raises(ValueError, match="yolcu|şablon|format",):
        services.import_gate_visa_files([("invalid.csv", b"foo,bar\nx,y\n")], replace=True, batch_id="bad")
    assert services.get_summary().passenger_count == 2

    package, _ = services.build_operation_package(ids=[0])
    with zipfile.ZipFile(io.BytesIO(package)) as archive:
        assert {"yolcular.xlsx", "yolcular.csv", "rapor.json"}.issubset(archive.namelist())

    ok, _, count = services.undo_import("batch-1")
    assert ok is True
    assert count == 0


def test_import_parse_does_not_hold_mutation_lock(monkeypatch, tmp_path):
    """Ayrıştırma adımı MUTATION_LOCK dışında çalışmalı.

    Aksi halde bozuk/büyük tek bir dosyanın ayrıştırması takılırsa (ör.
    openpyxl içinde asılı kalırsa), uygulamadaki TÜM diğer yazma işlemleri
    de sonsuza dek bloklanır — üretimde gözlemlenen 'dosyalar İşleniyor'da
    takılı kalıyor' hatasının kök nedeni buydu.
    """
    _isolate_store(monkeypatch, tmp_path)
    from backend import services

    parse_started = threading.Event()
    release_parse = threading.Event()
    original_parse = services._parse_import_files

    def slow_parse(files):
        parse_started.set()
        assert release_parse.wait(timeout=5), "test sinyali zaman aşımına uğradı"
        return original_parse(files)

    monkeypatch.setattr(services, "_parse_import_files", slow_parse)

    result_holder: dict = {}

    def run_import():
        result_holder["result"] = services.import_gate_visa_files(
            [("slow.csv", _csv("P999999", "2026-07-15"))],
            batch_id="slow-batch",
            dup_strategy="skip",
        )

    thread = threading.Thread(target=run_import)
    thread.start()
    try:
        assert parse_started.wait(timeout=5), "ayrıştırma başlamadı"

        # Ayrıştırma hâlâ sürerken başka bir kilitli mutasyon HEMEN tamamlanmalı.
        start = time.monotonic()
        assert services.delete_import_job("does-not-exist") is False
        elapsed = time.monotonic() - start
        assert elapsed < 1.0, "ayrıştırma MUTATION_LOCK'ı tutuyor gibi görünüyor"
    finally:
        release_parse.set()
        thread.join(timeout=5)
    assert not thread.is_alive()
    assert result_holder["result"][0] == 1


def test_import_parse_timeout_reports_error_without_hanging(monkeypatch, tmp_path):
    """Sonsuza dek dönmeyen bir ayrıştırma, iş kuyruğunu asmak yerine zaman
    aşımıyla hata vermeli."""
    _isolate_store(monkeypatch, tmp_path)
    from backend import services

    def never_returns(files):
        time.sleep(5)
        raise AssertionError("zaman aşımından önce dönmemeliydi")

    monkeypatch.setattr(services, "_parse_import_files", never_returns)

    with pytest.raises(TimeoutError):
        services._parse_import_files_with_timeout(
            [("slow.csv", _csv("P111111", "2026-07-15"))],
            timeout=0.2,
        )


def test_first_run_auth_roles_and_cookie(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "1")
    monkeypatch.setenv("APP_ENV", "development")

    from fastapi.testclient import TestClient
    from backend.auth import _LOGIN_FAILURES
    from backend.main import app

    _LOGIN_FAILURES.clear()
    with TestClient(app) as client:
        assert client.get("/api/auth/status").json()["setup_required"] is True
        assert client.get("/api/summary").status_code == 401

        setup = client.post("/api/auth/setup", json={"display_name": "Test Admin", "pin": "642975"})
        assert setup.status_code == 200
        assert setup.json()["user"]["role"] == "admin"
        assert client.get("/api/summary").status_code == 200

        upload = client.post(
            "/api/import?dup_strategy=skip&batch_id=api-batch",
            files=[
                ("files", ("one.csv", _csv("P111111", "2026-07-15"), "text/csv")),
                ("files", ("two.csv", _csv("P222222", "2026-07-16"), "text/csv")),
            ],
        )
        assert upload.status_code == 200
        assert upload.json()["imported"] == 2

        empty_upload = client.post(
            "/api/import?dup_strategy=skip&batch_id=empty-batch",
            files=[(
                "files",
                ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            )],
        )
        assert empty_upload.status_code == 400
        assert "0 bayt" in empty_upload.json()["detail"]

        photo = client.post(
            "/api/photos/match",
            files=[("files", ("P111111.jpg", b"\xff\xd8\xff\xe0synthetic", "image/jpeg"))],
        )
        assert photo.status_code == 200
        assert photo.json()["matched"] == 1
        assert client.get("/api/passengers").json()[0]["photo"]

        invalid_preview = client.post(
            "/api/import/preview",
            files={"file": ("invalid.csv", b"foo,bar\nx,y\n", "text/csv")},
        )
        assert invalid_preview.status_code == 400
        assert client.post("/api/import/undo?batch_id=api-batch").json()["passenger_count"] == 0

        viewer = client.post("/api/users", json={"name": "Viewer", "pin": "319764", "role": "viewer"})
        assert viewer.status_code == 200
        assert client.post("/api/auth/logout").status_code == 200
        assert client.get("/api/summary").status_code == 401

        assert client.post("/api/auth/login", json={"pin": "319764"}).status_code == 200
        forbidden = client.post("/api/passengers/clear")
        assert forbidden.status_code == 403


def test_eml_attachment_parser():
    from backend.mail_ingest import parse_eml

    message = EmailMessage()
    message["Subject"] = "PAX files"
    message["From"] = "ops@example.com"
    message.set_content("Attached")
    message.add_attachment(_csv("P111111", "2026-07-15"), maintype="text", subtype="csv", filename="pax.csv")

    parsed = parse_eml(message.as_bytes())
    assert parsed["subject"] == "PAX files"
    assert parsed["sender"] == "ops@example.com"
    assert parsed["attachments"][0]["filename"] == "pax.csv"


def test_bulk_list_count_limit_disabled_by_default():
    from backend.config import MAX_UPLOAD_FILES

    assert MAX_UPLOAD_FILES == int(os.environ.get("GATEVISA_MAX_UPLOAD_FILES", "0"))



@pytest.mark.parametrize(
    ("filename", "engine"),
    [
        ("pax.xlsx", "openpyxl"),
        ("pax.xlsm", "openpyxl"),
        ("pax.xls", "xlrd"),
        ("pax.ods", "odf"),
    ],
)
def test_excel_engine_is_selected_from_filename(filename, engine):
    from excelbase_core import excel_engine_for_filename

    assert excel_engine_for_filename(filename) == engine


def test_gate_visa_reader_passes_explicit_excel_engine(monkeypatch):
    import gate_visa_reader

    captured = {}

    class EmptyExcel:
        sheet_names = []

    def fake_excel_file(source, *, engine=None):
        captured["engine"] = engine
        return EmptyExcel()

    monkeypatch.setattr(gate_visa_reader.pd, "ExcelFile", fake_excel_file)
    with pytest.raises(ValueError, match="okunabilir yolcu verisi"):
        gate_visa_reader.read_gate_visa_file_bytes("22.05.xlsx", b"synthetic")

    assert captured["engine"] == "openpyxl"



def test_real_gate_visa_xlsx_is_parsed_with_openpyxl():
    from openpyxl import load_workbook
    from gate_visa_reader import build_gate_visa_template_xlsx, read_gate_visa_file_bytes

    workbook = load_workbook(io.BytesIO(build_gate_visa_template_xlsx()))
    sheet = workbook.active
    values = [1, "JANE", "DOE", "X1234567", "V-42", "2026-07-20", "2026-07-22", 25, 0]
    for column, value in enumerate(values, start=1):
        sheet.cell(row=5, column=column, value=value)
    payload = io.BytesIO()
    workbook.save(payload)

    results = read_gate_visa_file_bytes("22.05.xlsx", payload.getvalue())

    assert len(results) == 1
    assert results[0].rows == 1
    assert results[0].dataframe.iloc[0]["PASSPORT NUMBER"] == "X1234567"


def test_import_fails_loudly_when_db_write_fails(monkeypatch, tmp_path):
    """DB açıkken yazma başarısızsa import 'başarılı' dönmemeli (sessiz veri kaybı)."""
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")
    monkeypatch.setattr(db, "enabled", lambda: True)
    monkeypatch.setattr(db, "save_state", lambda payload: False)
    monkeypatch.setattr(db, "load_state", lambda: {})
    monkeypatch.setattr(db, "save_daily_backup", lambda payload: False)

    from fastapi.testclient import TestClient
    from backend.main import app

    with TestClient(app) as client:
        response = client.post(
            "/api/import?dup_strategy=skip&batch_id=db-fail",
            files=[("files", ("one.csv", _csv("P333333", "2026-07-15"), "text/csv"))],
        )
        assert response.status_code == 503
        assert "veritaban" in response.json()["detail"].lower()
        assert client.get("/api/passengers").json() == []
        assert client.get("/api/summary").json()["persistence"] == "database"


def test_summary_reports_local_fallback_persistence(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    from backend import services
    from backend.state import APP_VERSION

    summary = services.get_summary()
    assert summary.persistence == "local-fallback"
    assert summary.version == APP_VERSION


def test_ui_and_backend_versions_match():
    from pathlib import Path
    from backend.state import APP_VERSION

    version_ts = (Path(__file__).parents[1] / "frontend" / "src" / "lib" / "version.ts").read_text(encoding="utf-8")
    assert f'"{APP_VERSION}"' in version_ts


def test_cache_headers_keep_shell_and_api_fresh_but_cache_fingerprinted_assets(monkeypatch, tmp_path):
    """HTML/API taze kalırken içerik özetli statikler yeniden indirilmesin."""
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    import backend.main as backend_main
    from fastapi.testclient import TestClient

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    (out_dir / "index.html").write_text("<!doctype html><title>GV</title>", encoding="utf-8")
    monkeypatch.setattr(backend_main, "FRONTEND_OUT", out_dir)

    with TestClient(backend_main.app) as client:
        page = client.get("/")
        assert page.status_code == 200
        assert page.headers["cache-control"] == "no-store"

        api = client.get("/api/summary")
        assert api.status_code == 200
        assert api.headers["cache-control"] == "no-store"

        async def asset_response(_request):
            from fastapi.responses import Response

            return Response("console.log('ok')", media_type="text/javascript")

        import asyncio
        from starlette.requests import Request

        asset_request = Request(
            {
                "type": "http",
                "method": "GET",
                "path": "/_next/static/chunks/app.js",
                "headers": [],
                "query_string": b"",
                "scheme": "http",
                "server": ("testserver", 80),
                "client": ("testclient", 50000),
                "root_path": "",
            }
        )
        asset = asyncio.run(backend_main.cache_headers(asset_request, asset_response))
        assert asset.headers["cache-control"] == "public, max-age=31536000, immutable"

        health = client.get("/health")
        assert health.json()["version"]
        assert health.json()["database_writable"] is False


def _wait_queue_idle(client, timeout_seconds: float = 15.0) -> dict:
    import time as _time

    deadline = _time.monotonic() + timeout_seconds
    state = client.get("/api/import/queue").json()
    while state["active"] and _time.monotonic() < deadline:
        _time.sleep(0.1)
        state = client.get("/api/import/queue").json()
    return state


@pytest.mark.import_worker
def test_background_import_queue_processes_without_client(monkeypatch, tmp_path):
    """Dosyalar teslim edildikten sonra işleme istemciden bağımsız sürmeli."""
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend.main import app

    with TestClient(app) as client:
        enqueue = client.post(
            "/api/import/queue?dup_strategy=skip",
            files=[
                ("files", ("one.csv", _csv("P444444", "2026-07-15"), "text/csv")),
                ("files", ("two.csv", _csv("P555555", "2026-07-16"), "text/csv")),
                ("files", ("broken.xlsx", b"gecersiz icerik", "application/octet-stream")),
            ],
        )
        assert enqueue.status_code == 202
        assert len(enqueue.json()["jobs"]) == 3

        state = _wait_queue_idle(client)
        statuses = {job["filename"]: job["status"] for job in state["jobs"]}
        assert statuses["one.csv"] == "done"
        assert statuses["two.csv"] == "done"
        assert statuses["broken.xlsx"] == "error"
        assert state["active"] is False
        assert client.get("/api/summary").json()["passenger_count"] == 2

        # Hatalı iş yeniden kuyruğa alınabilmeli ve yine hata vermeli
        broken = next(job for job in state["jobs"] if job["filename"] == "broken.xlsx")
        retry = client.post(f"/api/import/queue/{broken['id']}/retry")
        assert retry.status_code == 200
        state = _wait_queue_idle(client)
        broken = next(job for job in state["jobs"] if job["filename"] == "broken.xlsx")
        assert broken["status"] == "error"

        # Kaldırma kuyruğu temizlemeli
        assert client.delete(f"/api/import/queue/{broken['id']}").status_code == 200
        names = [job["filename"] for job in client.get("/api/import/queue").json()["jobs"]]
        assert "broken.xlsx" not in names


@pytest.mark.import_worker
def test_zip_import_queue_expands_all_passenger_lists(monkeypatch, tmp_path):
    """Tek ZIP içindeki tüm desteklenen listeler ayrı kuyruk işi olmalı."""
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend.main import app

    archive_bytes = io.BytesIO()
    with zipfile.ZipFile(archive_bytes, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("one.csv", _csv("PZIP001", "2026-07-17", "ALICE"))
        archive.writestr("nested/two.csv", _csv("PZIP002", "2026-07-18", "BOB"))
        archive.writestr("__MACOSX/._one.csv", b"metadata")
        archive.writestr("notes.txt", b"ignored")

    with TestClient(app) as client:
        enqueue = client.post(
            "/api/import/queue?dup_strategy=skip&batch_id=zip-batch",
            files=[("files", ("49-listeler.zip", archive_bytes.getvalue(), "application/zip"))],
        )
        assert enqueue.status_code == 202
        assert [job["filename"] for job in enqueue.json()["jobs"]] == ["49-listeler.zip"]

        state = _wait_queue_idle(client)
        assert state["active"] is False
        children = [job for job in state["jobs"] if job["parent_id"]]
        assert {job["filename"] for job in children} == {"one.csv", "two.csv"}
        assert all(job["status"] == "done" for job in children)
        parent = next(job for job in state["jobs"] if job["kind"] == "upload")
        assert parent["status"] == "done"
        assert parent["processed_files"] == 2
        assert client.get("/api/summary").json()["passenger_count"] == 2


@pytest.mark.import_worker
def test_zip_import_marks_unsafe_member_without_blocking_intake(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend.main import app

    archive_bytes = io.BytesIO()
    with zipfile.ZipFile(archive_bytes, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("../escape.csv", _csv("PZIP003", "2026-07-19"))

    with TestClient(app) as client:
        response = client.post(
            "/api/import/queue",
            files=[("files", ("unsafe.zip", archive_bytes.getvalue(), "application/zip"))],
        )
        assert response.status_code == 202
        state = _wait_queue_idle(client)
        child = next((job for job in state["jobs"] if job["parent_id"]), None)
        assert child is not None, state
        assert child["status"] == "error"
        assert "güvensiz" in child["message"].lower()


@pytest.mark.import_worker
def test_background_import_replace_applies_only_to_first_file(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend.main import app

    with TestClient(app) as client:
        seeded = client.post(
            "/api/import?dup_strategy=add&batch_id=seed",
            files=[("files", ("seed.csv", _csv("P666666", "2026-07-14"), "text/csv"))],
        )
        assert seeded.status_code == 200
        assert client.get("/api/summary").json()["passenger_count"] == 1

        enqueue = client.post(
            "/api/import/queue?replace=true&dup_strategy=skip",
            files=[
                ("files", ("r1.csv", _csv("P777777", "2026-07-15"), "text/csv")),
                ("files", ("r2.csv", _csv("P888888", "2026-07-16"), "text/csv")),
            ],
        )
        assert enqueue.status_code == 202
        _wait_queue_idle(client)
        # İlk dosya listeyi değiştirdi, ikincisi eklendi: eski kayıt gitti.
        assert client.get("/api/summary").json()["passenger_count"] == 2


@pytest.mark.import_worker
def test_zip_with_49_real_xlsx_is_accepted_as_one_job_then_processed(monkeypatch, tmp_path):
    """49 gerçek XLSX, HTTP isteği içinde açılmadan tek parent olarak teslim edilir."""
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend.main import app

    archive_bytes = io.BytesIO()
    with zipfile.ZipFile(archive_bytes, "w", zipfile.ZIP_DEFLATED) as archive:
        for index in range(49):
            archive.writestr(
                f"seferler/{index + 1:02d}.xlsx",
                _xlsx(f"X49{index:05d}", f"2026-09-{(index % 28) + 1:02d}", f"PAX{index}"),
            )

    with TestClient(app) as client:
        response = client.post(
            "/api/import/queue?dup_strategy=skip&upload_id=realistic-49-zip",
            files=[("files", ("49-listeler.zip", archive_bytes.getvalue(), "application/zip"))],
            headers={"X-Request-ID": "test-realistic-49"},
        )
        assert response.status_code == 202
        assert response.headers["x-request-id"] == "test-realistic-49"
        assert len(response.json()["jobs"]) == 1
        assert response.json()["jobs"][0]["filename"] == "49-listeler.zip"

        state = _wait_queue_idle(client, timeout_seconds=60)
        children = [job for job in state["jobs"] if job["parent_id"]]
        assert len(children) == 49
        assert all(job["status"] == "done" for job in children)
        parent = next(job for job in state["jobs"] if job["id"] == "realistic-49-zip")
        assert parent["status"] == "done"
        assert parent["processed_files"] == 49
        assert client.get("/api/summary").json()["passenger_count"] == 49


def test_zip_intake_does_not_expand_or_read_passenger_state_and_is_idempotent(monkeypatch, tmp_path):
    """İstek sadece bir ZIP kaydı yazar; 49 üye ve yavaş audit yanıtı geciktirmez."""
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend import services
    from backend.main import app

    archive_bytes = io.BytesIO()
    with zipfile.ZipFile(archive_bytes, "w", zipfile.ZIP_DEFLATED) as archive:
        for index in range(49):
            archive.writestr(f"{index:02d}.xlsx", _xlsx(f"FAST{index:04d}", "2026-10-01"))

    with TestClient(app) as client:
        original_enqueue = services._queue_enqueue_job
        calls = {"persist": 0}

        def one_slow_persist(*args, **kwargs):
            calls["persist"] += 1
            time.sleep(0.03)  # DB benzeri tek kalıcı yazma gecikmesi
            return original_enqueue(*args, **kwargs)

        monkeypatch.setattr(services, "_queue_enqueue_job", one_slow_persist)
        monkeypatch.setattr(services, "ensure_import_worker", lambda: None)
        monkeypatch.setattr(
            services,
            "expand_import_upload",
            lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("intake ZIP açmamalı")),
        )
        monkeypatch.setattr(
            services,
            "load_state",
            lambda: (_ for _ in ()).throw(AssertionError("intake yolcu state okumamalı")),
        )
        monkeypatch.setattr(services, "record_audit", lambda *args: time.sleep(0.5))

        started = time.monotonic()
        first = client.post(
            "/api/import/queue?upload_id=safe-mobile-retry&upload_index=7",
            files=[("files", ("all.zip", archive_bytes.getvalue(), "application/zip"))],
        )
        elapsed = time.monotonic() - started
        second = client.post(
            "/api/import/queue?upload_id=safe-mobile-retry&upload_index=7",
            files=[("files", ("all.zip", archive_bytes.getvalue(), "application/zip"))],
        )

        assert first.status_code == 202
        assert second.status_code == 202
        assert first.json()["jobs"][0]["id"] == second.json()["jobs"][0]["id"]
        assert elapsed < 0.4
        # İki HTTP denemesi var; fakat ikisi de aynı top-level idempotency kaydını
        # döndürür. ZIP'in 49 üyesi request thread'inde persist edilmez.
        assert calls["persist"] == 2
        assert len(services._queue_list()) == 1
        assert services._queue_get("safe-mobile-retry")["ordinal"] == 7


def test_configured_database_outage_returns_503_without_local_queue_fallback(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")
    monkeypatch.setenv("DATABASE_URL", "postgresql://temporarily-unavailable/example")

    from fastapi.testclient import TestClient
    from backend.main import app

    with TestClient(app) as client:
        response = client.post(
            "/api/import/queue?upload_id=must-stay-durable",
            files=[("files", ("one.csv", _csv("NODB001", "2026-10-02"), "text/csv"))],
        )
    assert response.status_code == 503
    assert "veritaban" in response.json()["detail"].lower()
    assert not (tmp_path / "import-queue" / "jobs.json").exists()


@pytest.mark.import_worker
def test_mixed_zip_keeps_good_files_and_reports_bad_member(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend.main import app

    archive_bytes = io.BytesIO()
    with zipfile.ZipFile(archive_bytes, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("good-one.xlsx", _xlsx("MIXED001", "2026-11-01", "ALICE"))
        archive.writestr("broken.xlsx", b"not an excel workbook")
        archive.writestr("nested/good-two.xlsx", _xlsx("MIXED002", "2026-11-02", "BOB"))

    with TestClient(app) as client:
        response = client.post(
            "/api/import/queue?upload_id=mixed-archive",
            files=[("files", ("mixed.zip", archive_bytes.getvalue(), "application/zip"))],
        )
        assert response.status_code == 202
        state = _wait_queue_idle(client, timeout_seconds=30)
        children = {job["filename"]: job for job in state["jobs"] if job["parent_id"]}
        assert children["good-one.xlsx"]["status"] == "done"
        assert children["good-two.xlsx"]["status"] == "done"
        assert children["broken.xlsx"]["status"] == "error"
        parent = next(job for job in state["jobs"] if job["id"] == "mixed-archive")
        assert parent["status"] == "done"
        assert parent["processed_files"] == 3
        assert client.get("/api/summary").json()["passenger_count"] == 2


@pytest.mark.import_worker
def test_bad_first_zip_member_does_not_consume_replace(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend.main import app

    archive_bytes = io.BytesIO()
    with zipfile.ZipFile(archive_bytes, "w", zipfile.ZIP_DEFLATED) as archive:
        # Boyutu sıfır değil: child oluşturulur ve ancak gerçek Excel parse
        # aşamasında bozuk olduğu anlaşılır. Replace hakkını tüketmemeli.
        archive.writestr("00-broken.xlsx", b"this is not an xlsx workbook")
        archive.writestr("01-valid.xlsx", _xlsx("REPLACE2", "2026-12-02", "NEW"))

    with TestClient(app) as client:
        seeded = client.post(
            "/api/import?dup_strategy=add&batch_id=seed-before-replace",
            files=[("files", ("seed.csv", _csv("REPLACE1", "2026-12-01", "OLD"), "text/csv"))],
        )
        assert seeded.status_code == 200
        response = client.post(
            "/api/import/queue?replace=true&upload_id=replace-after-bad",
            files=[("files", ("replace.zip", archive_bytes.getvalue(), "application/zip"))],
        )
        assert response.status_code == 202
        state = _wait_queue_idle(client, timeout_seconds=30)
        children = {job["filename"]: job for job in state["jobs"] if job["parent_id"]}
        assert children["00-broken.xlsx"]["status"] == "error"
        assert children["01-valid.xlsx"]["status"] == "done"
        passengers = client.get("/api/passengers").json()
        assert len(passengers) == 1
        assert passengers[0]["passport_no"] == "REPLACE2"


@pytest.mark.import_worker
def test_bad_first_top_level_file_does_not_consume_batch_replace(monkeypatch, tmp_path):
    """Ayrı mobil POST'larda da replace ilk 202'de değil ilk başarılı parse'ta tüketilir."""
    _isolate_store(monkeypatch, tmp_path)
    monkeypatch.setenv("GATEVISA_REQUIRE_AUTH", "0")

    from fastapi.testclient import TestClient
    from backend.main import app

    with TestClient(app) as client:
        seeded = client.post(
            "/api/import?dup_strategy=add&batch_id=top-level-seed",
            files=[("files", ("seed.csv", _csv("TOPOLD1", "2026-12-04", "OLD"), "text/csv"))],
        )
        assert seeded.status_code == 200

        batch_id = "top-level-replace-batch"
        broken = client.post(
            f"/api/import/queue?replace=true&batch_id={batch_id}&upload_id=top-bad&upload_index=0",
            files=[("files", ("00-broken.xlsx", b"not a workbook", "application/octet-stream"))],
        )
        valid = client.post(
            f"/api/import/queue?replace=true&batch_id={batch_id}&upload_id=top-good&upload_index=1",
            files=[("files", ("01-valid.xlsx", _xlsx("TOPNEW2", "2026-12-05", "NEW"), "application/octet-stream"))],
        )
        assert broken.status_code == 202
        assert valid.status_code == 202

        state = _wait_queue_idle(client, timeout_seconds=30)
        top_jobs = {job["filename"]: job for job in state["jobs"] if not job["parent_id"]}
        assert top_jobs["00-broken.xlsx"]["status"] == "error"
        assert top_jobs["01-valid.xlsx"]["status"] == "done"
        passengers = client.get("/api/passengers").json()
        assert [passenger["passport_no"] for passenger in passengers] == ["TOPNEW2"]


@pytest.mark.import_worker
def test_local_queue_recovers_processing_job_after_restart(monkeypatch, tmp_path):
    _isolate_store(monkeypatch, tmp_path)
    from backend import services

    jobs, _ = services.enqueue_import_uploads(
        [("restart.csv", _csv("RESTART1", "2026-12-03"), "text/csv")],
        upload_id="restart-job",
    )
    assert jobs[0]["status"] == "pending"
    claimed = services._queue_claim()
    assert claimed and claimed["status"] == "processing"

    assert services.recover_stale_import_jobs() == 1
    assert services._queue_get("restart-job")["status"] == "pending"
    services.ensure_import_worker()
    deadline = time.monotonic() + 10
    while services.get_import_jobs()[1] and time.monotonic() < deadline:
        time.sleep(0.05)
    assert services._queue_get("restart-job")["status"] == "done"
    assert services.get_summary().passenger_count == 1


@pytest.mark.import_worker
def test_recovered_add_job_is_idempotent_if_state_saved_before_worker_crash(monkeypatch, tmp_path):
    """At-least-once lease, add stratejisinde aynı yolcuyu iki kez yazmamalı."""
    _isolate_store(monkeypatch, tmp_path)
    from backend import services

    jobs, batch_id = services.enqueue_import_uploads(
        [("once.csv", _csv("ONCE001", "2026-12-04"), "text/csv")],
        dup_strategy="add",
        upload_id="crash-after-save",
    )
    claimed = services._queue_claim()
    assert claimed is not None
    # Yolcu state kaydı başarıyla bitti, fakat queue finish çağrısından önce
    # process öldü senaryosu.
    result = services.import_gate_visa_files(
        [("once.csv", claimed["payload"])],
        dup_strategy="add",
        batch_id=batch_id,
        job_id=jobs[0]["id"],
    )
    assert result[0] == 1
    assert services.get_summary().passenger_count == 1

    assert services.recover_stale_import_jobs(force=True) == 1
    services.ensure_import_worker()
    deadline = time.monotonic() + 10
    while services.get_import_jobs()[1] and time.monotonic() < deadline:
        time.sleep(0.05)
    assert services._queue_get("crash-after-save")["status"] == "done"
    assert services.get_summary().passenger_count == 1


def test_daily_backup_is_throttled(monkeypatch, tmp_path):
    """Art arda kayıtlar tüm veriyi her seferinde yeniden yedeklemesin."""
    import pandas as pd

    calls = {"backup": 0}

    def fake_backup(payload):
        calls["backup"] += 1
        return True

    df = pd.DataFrame(columns=persistence.ALL_COLUMNS)
    # NOT: _last_backup_at'i 0.0'a sıfırlamak YANLIŞTI — time.monotonic()'in
    # mutlak taban değeri platforma göre değişir (taze bir CI konteynerinde
    # birkaç saniye, uzun süredir çalışan bir makinede binlerce saniye
    # olabilir); "now - 0.0 >= 600" karşılaştırması konteyner henüz 600 sn
    # ayakta değilse yanlışlıkla False dönüp testi CI'da ara sıra
    # başarısız kılıyordu (yerelde hep geçiyordu çünkü sandbox'ın monotonic
    # tabanı zaten binlerce saniyeydi). -inf, mutlak tabandan tamamen
    # bağımsız olarak "süre her zaman aşıldı" anlamına gelir.
    with persistence._STORE_LOCK:
        monkeypatch.setattr(db, "enabled", lambda: True)
        monkeypatch.setattr(db, "save_state", lambda payload: True)
        monkeypatch.setattr(db, "save_daily_backup", fake_backup)
        monkeypatch.setattr(persistence, "_last_backup_at", float("-inf"))
        persistence.save_store(df, [], {})
        persistence.save_store(df, [], {})
        persistence.save_store(df, [], {})
        observed = calls["backup"]
    assert observed == 1


def test_db_engine_retries_after_transient_failure(monkeypatch, tmp_path):
    """Açılışta DB'ye ulaşılamazsa bağlantı kalıcı olarak kapanmamalı."""
    calls = {"n": 0}
    real_create_engine = db.create_engine

    def flaky_create_engine(url, **kwargs):
        calls["n"] += 1
        if calls["n"] == 1:
            raise RuntimeError("bağlantı reddedildi")
        return real_create_engine(url, **kwargs)

    monkeypatch.setattr(db, "create_engine", flaky_create_engine)
    monkeypatch.setattr(db, "_engine", None)
    monkeypatch.setattr(db, "_init_done", False)
    monkeypatch.setattr(db, "_init_failed", False)
    monkeypatch.setattr(db, "_retry_at", 0.0)
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path}/retry.db")

    assert db.get_engine() is None  # ilk deneme başarısız
    assert db.get_engine() is None  # bekleme süresi dolmadan yeniden denenmez
    assert calls["n"] == 1

    monkeypatch.setattr(db, "_retry_at", 0.0)  # bekleme süresi doldu
    assert db.get_engine() is not None
    assert calls["n"] == 2
