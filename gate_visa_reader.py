from __future__ import annotations

import io
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from excelbase_core import ReadResult, format_cell_value, is_blank

TEMPLATE_TITLE = "GATE VISA PAX LIST"

# Excel şablondaki sabit kolon anahtarları (satır 3-4 başlıklarından üretilir)
EXCEL_COLUMNS = [
    "NO",
    "NAME",
    "SURNAME",
    "PASSPORT NUMBER",
    "VOUCHER",
    "DEPARTURE",
    "ARRIVAL",
    "ADULT",
    "CHILD",
]


def norm_cell(value: Any) -> str:
    text = str(value if value is not None else "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def is_gate_visa_raw(raw: pd.DataFrame) -> bool:
    """Returns True if the sheet looks like a Gate Visa / passenger-list Excel."""
    flat = " ".join(norm_cell(v) for row in raw.head(10).itertuples(index=False) for v in row)
    if "gate visa" in flat or "pax list" in flat:
        return True
    # Look wider (up to row 20) and be permissive: 2 out of 3 key words is enough
    for i in range(min(20, len(raw))):
        cells = [norm_cell(v) for v in raw.iloc[i].tolist()]
        has_name = "name" in cells
        has_surname = "surname" in cells
        has_passport = any("passport" in c for c in cells)
        has_voucher = any("voucher" in c for c in cells)
        has_no = "no" in cells or "#" in cells
        if (has_name and has_surname) or (has_name and has_passport) or (has_surname and has_passport):
            return True
        if has_name and has_voucher and has_no:
            return True
    return False


def find_header_rows(raw: pd.DataFrame) -> tuple[int, int | None]:
    for i in range(min(12, len(raw))):
        cells = [norm_cell(v) for v in raw.iloc[i].tolist()]
        if "name" not in cells or "surname" not in cells:
            continue
        if i + 1 < len(raw):
            next_cells = [norm_cell(v) for v in raw.iloc[i + 1].tolist()]
            if any(x in next_cells for x in ("departure", "arrival", "adult", "child")):
                return i, i + 1
        return i, None
    return -1, None


def build_column_names(main_row: pd.Series, sub_row: pd.Series | None) -> list[str]:
    names: list[str] = []
    sub_values = sub_row.tolist() if sub_row is not None else [None] * len(main_row)

    for main, sub in zip(main_row.tolist(), sub_values):
        main_n = norm_cell(main)
        sub_n = norm_cell(sub)

        if main_n in ("no", "#"):
            names.append("NO")
        elif main_n == "name":
            names.append("NAME")
        elif main_n == "surname":
            names.append("SURNAME")
        elif "passport" in main_n:
            names.append("PASSPORT NUMBER")
        elif main_n == "voucher":
            names.append("VOUCHER")
        elif sub_n == "departure":
            names.append("DEPARTURE")
        elif sub_n == "arrival":
            names.append("ARRIVAL")
        elif sub_n == "adult":
            names.append("ADULT")
        elif sub_n == "child":
            names.append("CHILD")
        elif main_n in ("departure", "arrival", "adult", "child"):
            names.append(main_n.upper())
        elif main_n:
            names.append(main_n.upper())
        elif sub_n:
            names.append(sub_n.upper())
        else:
            names.append(f"COL_{len(names) + 1}")
    return names


def align_to_schema(columns: list[str]) -> list[str]:
    """Şablondaki 9 kolona oturt; fazla kolonları at, eksikleri doldur."""
    mapping: dict[str, str] = {}
    for col in columns:
        key = norm_cell(col).replace("_", " ")
        if key in ("no", "#"):
            mapping["NO"] = col
        elif key == "name":
            mapping["NAME"] = col
        elif key == "surname":
            mapping["SURNAME"] = col
        elif "passport" in key:
            mapping["PASSPORT NUMBER"] = col
        elif key == "voucher":
            mapping["VOUCHER"] = col
        elif key == "departure":
            mapping["DEPARTURE"] = col
        elif key == "arrival":
            mapping["ARRIVAL"] = col
        elif key == "adult":
            mapping["ADULT"] = col
        elif key == "child":
            mapping["CHILD"] = col

    return [mapping.get(field, field) for field in EXCEL_COLUMNS]


def parse_gate_visa_raw(raw: pd.DataFrame) -> pd.DataFrame:
    raw = raw.dropna(how="all").dropna(axis=1, how="all")
    if raw.empty:
        return pd.DataFrame(columns=EXCEL_COLUMNS)

    main_idx, sub_idx = find_header_rows(raw)
    if main_idx < 0:
        raise ValueError("Gate Visa şablonu bulunamadı. NAME / SURNAME başlıkları gerekli.")

    main_row = raw.iloc[main_idx]
    sub_row = raw.iloc[sub_idx] if sub_idx is not None else None
    col_names = build_column_names(main_row, sub_row)

    data_start = (sub_idx if sub_idx is not None else main_idx) + 1
    data = raw.iloc[data_start:].copy()
    data.columns = col_names[: len(data.columns)]
    data = data.dropna(how="all")

    out = pd.DataFrame()
    for field in EXCEL_COLUMNS:
        source_col = next((c for c in data.columns if norm_cell(c).replace("_", " ") == norm_cell(field)), None)
        if source_col is None and field in data.columns:
            source_col = field
        if source_col and source_col in data.columns:
            out[field] = data[source_col].map(format_cell_value)
        else:
            out[field] = ""

    # Tamamen boş satırları at
    out = out[~out.apply(lambda row: all(is_blank(v) for v in row), axis=1)]
    return out.reset_index(drop=True)


def read_gate_visa_sheet(excel: pd.ExcelFile, sheet: str) -> pd.DataFrame:
    raw = pd.read_excel(excel, sheet_name=sheet, header=None, dtype=object)
    if not is_gate_visa_raw(raw):
        raise ValueError("Bu sayfa Gate Visa PAX şablonu değil.")
    return parse_gate_visa_raw(raw)


def _fallback_parse(sheet_df: pd.DataFrame, file_name: str, sheet_name: str) -> "pd.DataFrame | None":
    """Try to extract passenger data from any Excel that has recognisable columns."""
    from excelbase_core import detect_header_row_index, make_unique_columns, format_cell_value, is_blank
    sheet_df = sheet_df.dropna(how="all").dropna(axis=1, how="all")
    if sheet_df.empty:
        return None
    hdr_idx = detect_header_row_index(sheet_df)
    raw_headers = sheet_df.iloc[hdr_idx].tolist()
    data = sheet_df.iloc[hdr_idx + 1:].copy()
    if data.empty:
        return None
    data.columns = make_unique_columns(raw_headers)
    data = data.dropna(how="all")
    for col in data.columns:
        data[col] = data[col].map(format_cell_value)
    data = data[~data.apply(lambda r: all(is_blank(v) for v in r), axis=1)]
    if data.empty:
        return None

    EXCEL_COL_MAP = {
        "NO": ["no", "#", "sira", "sıra"],
        "NAME": ["name", "ad", "isim", "first name", "firstname"],
        "SURNAME": ["surname", "soyad", "soyisim", "last name", "lastname"],
        "PASSPORT NUMBER": ["passport", "pasaport", "passport number", "passport no", "doc no"],
        "VOUCHER": ["voucher", "pnr", "bilet", "ticket", "reservation"],
        "DEPARTURE": ["departure", "gidis", "gidiş", "depart"],
        "ARRIVAL": ["arrival", "varis", "varış", "arrive"],
        "ADULT": ["adult", "yetiskin", "yetişkin", "adult fee"],
        "CHILD": ["child", "cocuk", "çocuk", "child fee"],
    }

    def find_col(candidates: list[str]) -> "str | None":
        for col in data.columns:
            col_n = norm_cell(col).replace(" ", "")
            for cand in candidates:
                cand_n = norm_cell(cand).replace(" ", "")
                if cand_n and (cand_n in col_n or col_n in cand_n):
                    return col
        return None

    out = pd.DataFrame()
    for field, syns in EXCEL_COL_MAP.items():
        src = find_col(syns)
        out[field] = data[src].astype(str) if src and src in data.columns else ""
    out = out[~out.apply(lambda r: all(is_blank(v) for v in r), axis=1)]
    return out.reset_index(drop=True) if not out.empty else None


def read_gate_visa_file_bytes(file_name: str, raw_bytes: bytes) -> list[ReadResult]:
    lower = file_name.lower()
    if lower.endswith(".csv"):
        text = raw_bytes.decode("utf-8-sig", errors="replace")
        df_raw = pd.read_csv(io.StringIO(text), header=None, dtype=object)
        if is_gate_visa_raw(df_raw):
            df = parse_gate_visa_raw(df_raw)
        else:
            fallback = _fallback_parse(df_raw, file_name, "CSV")
            if fallback is None or fallback.empty:
                raise ValueError("CSV Gate Visa şablonu formatında değil.")
            df = fallback
        return [ReadResult(file_name, "CSV", len(df), len(df.columns), df)]

    if not lower.endswith((".xlsx", ".xls", ".xlsm", ".ods")):
        raise ValueError("Desteklenen dosya türleri: .xlsx, .xls, .xlsm, .ods, .csv")

    excel = pd.ExcelFile(io.BytesIO(raw_bytes))
    results: list[ReadResult] = []
    for sheet in excel.sheet_names:
        try:
            sheet_df = pd.read_excel(excel, sheet_name=sheet, header=None, dtype=object)
            if is_gate_visa_raw(sheet_df):
                df = parse_gate_visa_raw(sheet_df)
            else:
                fallback = _fallback_parse(sheet_df, file_name, str(sheet))
                if fallback is None or fallback.empty:
                    continue
                df = fallback
            if len(df) > 0:
                results.append(ReadResult(file_name, str(sheet), len(df), len(df.columns), df))
        except Exception as exc:
            results.append(
                ReadResult(file_name, str(sheet), 0, 0, pd.DataFrame({"Hata": [f"Sayfa okunamadı: {exc}"]}))
            )

    if not results:
        raise ValueError(
            "Excel'de okunabilir yolcu verisi bulunamadı.\n"
            "Beklenen format: GATE VISA PAX LIST (NAME / SURNAME / PASSPORT NUMBER kolonları)."
        )
    return results


def build_gate_visa_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PAX LIST"

    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="FFD966")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header_font = Font(bold=True, size=10)

    ws.merge_cells("A1:I1")
    ws["A1"] = TEMPLATE_TITLE
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, label in enumerate(["NO", "NAME", "SURNAME", "PASSPORT NUMBER", "VOUCHER"], start=1):
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("F3:G3")
    ws["F3"] = "DATE"
    ws["F3"].fill = header_fill
    ws["F3"].font = header_font
    ws["F3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F4"] = "DEPARTURE"
    ws["G4"] = "ARRIVAL"
    for col in (6, 7):
        c = ws.cell(row=4, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("H3:I3")
    ws["H3"] = "VISA FEE"
    ws["H3"].fill = header_fill
    ws["H3"].font = header_font
    ws["H3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["H4"] = "ADULT"
    ws["I4"] = "CHILD"
    for col in (8, 9):
        c = ws.cell(row=4, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    widths = [6, 14, 14, 18, 14, 14, 14, 10, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
