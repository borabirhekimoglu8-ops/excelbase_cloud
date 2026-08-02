"""Reads a work folder and works out what the application is missing.

No model is involved, deliberately. The questions this answers -- which
spreadsheet templates recur, which columns appear in the files but nowhere in
the app, how much of the folder is documents versus tables -- have exact
answers that can be counted. Sending the folder to a language model would
trade an exact answer for a plausible one, cost money per run, differ between
runs, and carry passport scans off the device to do it.

What it deliberately cannot do: read prose. A PDF of a procedure is counted
and named, never opened. Anything claiming to understand the *content* of the
documents would have to send them somewhere.

Only the header row of a spreadsheet is read. That is what identifies a
template, it keeps a scan of thousands of files to seconds, and it keeps
passenger rows out of the report entirely.
"""

from __future__ import annotations

import csv
import logging
import os
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)

SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm", ".xls", ".csv", ".ods"}
DOCUMENT_SUFFIXES = {".pdf", ".doc", ".docx", ".rtf", ".txt"}
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".heic", ".webp"}

# A scan must finish while someone is waiting, and a runaway folder must not
# take the server down with it.
MAX_FILES = 20_000
MAX_SPREADSHEETS_READ = 1_500
MAX_HEADER_COLUMNS = 60

# Noise in every Windows/Drive tree.
SKIP_DIRECTORIES = {
    ".git", "node_modules", "__pycache__", ".venv", "venv",
    "$RECYCLE.BIN", "System Volume Information", ".tmp.drivedownload",
    ".Trash",
}

_DATE_IN_NAME = re.compile(r"20\d{2}[-_.]?(?:0[1-9]|1[0-2])[-_.]?(?:0[1-9]|[12]\d|3[01])")

# Columns the application already understands. Anything outside this is a gap
# between what the operator records and what the app can hold.
KNOWN_COLUMNS: frozenset[str] = frozenset({
    "no", "ad", "adi", "soyad", "soyadi", "ad soyad", "adsoyad",
    "pasaport", "pasaport no", "pasaportno", "passport", "passport no",
    "voucher", "rezervasyon", "rezervasyon no",
    "gidis", "gidis tarihi", "donus", "donus tarihi", "varis", "varis tarihi",
    "tarih", "ucret", "vize ucreti", "yetiskin", "cocuk",
    "durum", "fotograf", "evrak", "belge", "aciklama", "not",
})


@dataclass(frozen=True, slots=True)
class TemplateGroup:
    """Files that share a header row: the same form filled repeatedly."""

    headers: tuple[str, ...]
    files: tuple[str, ...]

    @property
    def count(self) -> int:
        return len(self.files)


@dataclass(slots=True)
class Finding:
    """One thing worth doing, with the evidence that suggested it."""

    kind: str
    title: str
    detail: str
    evidence: list[str] = field(default_factory=list)
    #: Ready to paste into the development panel. The operator approves this
    #: sentence; the folder's contents never travel with it.
    suggestion: str = ""
    #: Higher sorts first. Derived from how many files back the finding.
    weight: int = 0

    def as_dict(self) -> dict:
        return {
            "kind": self.kind,
            "title": self.title,
            "detail": self.detail,
            "evidence": list(self.evidence),
            "suggestion": self.suggestion,
            "weight": self.weight,
        }


@dataclass(slots=True)
class ScanReport:
    root: str
    files_seen: int
    truncated: bool
    by_kind: dict[str, int]
    templates: list[TemplateGroup]
    unknown_columns: list[tuple[str, int]]
    dated_folders: int
    findings: list[Finding]

    def as_dict(self) -> dict:
        return {
            "root": self.root,
            "files_seen": self.files_seen,
            "truncated": self.truncated,
            "by_kind": dict(self.by_kind),
            "templates": [
                {"headers": list(group.headers), "count": group.count, "files": list(group.files[:5])}
                for group in self.templates
            ],
            "unknown_columns": [{"column": name, "files": count} for name, count in self.unknown_columns],
            "dated_folders": self.dated_folders,
            "findings": [finding.as_dict() for finding in self.findings],
        }


def normalize_column(value: str) -> str:
    """Fold a header to a comparable key.

    Turkish exports vary the same column endlessly -- "Pasaport No", "PASAPORT
    NO", "Pasaport_No", "Pasaport  Nº". Without folding, one column looks like
    four and every count is wrong.
    """
    text = unicodedata.normalize("NFKD", value).strip().lower()
    replacements = {"ı": "i", "İ": "i", "ş": "s", "ğ": "g", "ü": "u", "ö": "o", "ç": "c"}
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _kind_for(suffix: str) -> str:
    lowered = suffix.lower()
    if lowered in SPREADSHEET_SUFFIXES:
        return "tablo"
    if lowered in DOCUMENT_SUFFIXES:
        return "belge"
    if lowered in IMAGE_SUFFIXES:
        return "gorsel"
    return "diger"


def _csv_headers(path: Path) -> list[str]:
    # utf-8-sig because Excel writes a BOM, which would otherwise become part
    # of the first column's name and make it look like a column of its own.
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        for row in csv.reader(handle, dialect):
            if any(cell.strip() for cell in row):
                return row[:MAX_HEADER_COLUMNS]
    return []


def _excel_headers(path: Path) -> list[str]:
    from openpyxl import load_workbook

    # read_only + values_only streams a single row instead of loading the
    # workbook; a 50MB export otherwise costs seconds and hundreds of MB each.
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            return []
        for row in sheet.iter_rows(min_row=1, max_row=8, values_only=True):
            cells = ["" if cell is None else str(cell).strip() for cell in row]
            if sum(1 for cell in cells if cell) >= 2:
                return cells[:MAX_HEADER_COLUMNS]
        return []
    finally:
        workbook.close()


def read_headers(path: Path) -> list[str]:
    """The header row of one spreadsheet, or [] when it cannot be read.

    A folder of real work always contains a locked, corrupt or half-synced
    file. One of those must not end the scan.
    """
    try:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return [cell.strip() for cell in _csv_headers(path)]
        if suffix in {".xlsx", ".xlsm"}:
            return [cell.strip() for cell in _excel_headers(path)]
        # .xls and .ods need heavier readers; they are counted, not parsed.
        return []
    except Exception:
        logger.info("başlık okunamadı: %s", path.name)
        return []


def _walk(root: Path) -> tuple[list[Path], bool, int]:
    # os.walk rather than Path.walk: the latter arrived in Python 3.12 and this
    # has to run on whatever interpreter the operator's venv was built with.
    files: list[Path] = []
    dated_folders = 0
    for current, directories, names in os.walk(root, onerror=lambda _error: None):
        directories[:] = [
            name for name in directories
            if name not in SKIP_DIRECTORIES and not name.startswith("~$")
        ]
        if _DATE_IN_NAME.search(Path(current).name):
            dated_folders += 1
        for name in names:
            if name.startswith("~$") or name.startswith("."):
                continue
            files.append(Path(current) / name)
            if len(files) >= MAX_FILES:
                return files, True, dated_folders
    return files, False, dated_folders


def scan_folder(root: str | Path) -> ScanReport:
    """Inventory a work folder and derive what the application should gain."""
    base = Path(root).expanduser()
    if not base.is_dir():
        raise NotADirectoryError(f"Klasör bulunamadı: {base}")

    files, truncated, dated_folders = _walk(base)
    by_kind: Counter[str] = Counter()
    spreadsheets: list[Path] = []
    for path in files:
        kind = _kind_for(path.suffix)
        by_kind[kind] += 1
        if kind == "tablo":
            spreadsheets.append(path)

    # Newest first: a template still in use matters more than one retired
    # years ago, and the cap has to fall on the stale end.
    spreadsheets.sort(key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True)

    grouped: dict[tuple[str, ...], list[str]] = {}
    column_files: Counter[str] = Counter()
    original_case: dict[str, str] = {}
    for path in spreadsheets[:MAX_SPREADSHEETS_READ]:
        headers = [cell for cell in read_headers(path) if cell]
        if len(headers) < 2:
            continue
        keys = tuple(normalize_column(cell) for cell in headers)
        grouped.setdefault(keys, []).append(str(path.relative_to(base)))
        for key, raw in zip(keys, headers):
            if not key:
                continue
            column_files[key] += 1
            original_case.setdefault(key, raw)

    templates = sorted(
        (TemplateGroup(headers=keys, files=tuple(paths)) for keys, paths in grouped.items()),
        key=lambda group: group.count,
        reverse=True,
    )
    unknown = sorted(
        ((original_case.get(key, key), count) for key, count in column_files.items()
         if key and key not in KNOWN_COLUMNS),
        key=lambda item: item[1],
        reverse=True,
    )

    return ScanReport(
        root=str(base),
        files_seen=len(files),
        truncated=truncated,
        by_kind=dict(by_kind),
        templates=templates,
        unknown_columns=unknown,
        dated_folders=dated_folders,
        findings=build_findings(templates, unknown, by_kind, dated_folders),
    )


def build_findings(
    templates: list[TemplateGroup],
    unknown_columns: list[tuple[str, int]],
    by_kind: dict[str, int],
    dated_folders: int,
) -> list[Finding]:
    """Turn counts into things worth building, each carrying its evidence.

    Every suggestion is a sentence the operator can hand to the development
    panel unchanged. None of them contains a file's contents -- only column
    names and counts -- so approving one does not send the folder anywhere.
    """
    findings: list[Finding] = []

    for group in templates[:5]:
        if group.count < 3:
            continue
        columns = ", ".join(header for header in group.headers if header)[:300]
        findings.append(Finding(
            kind="template",
            title=f"{group.count} dosya aynı şablonu kullanıyor",
            detail=(
                "Aynı sütun düzeni tekrar tekrar dolduruluyor. Uygulamada bu şablon için "
                "bir ekran olsaydı dosya açmadan girilebilirdi."
            ),
            evidence=[f"Sütunlar: {columns}", *[f"Örnek: {name}" for name in group.files[:3]]],
            suggestion=(
                f"Uygulamaya şu sütunlardan oluşan bir kayıt ekranı ve listesi ekle: {columns}. "
                "Excel ile içe aktarma ve dışa aktarma da olsun."
            ),
            weight=group.count * 10,
        ))

    strong_unknown = [(name, count) for name, count in unknown_columns if count >= 3][:12]
    if strong_unknown:
        listed = ", ".join(f"{name} ({count} dosya)" for name, count in strong_unknown)
        findings.append(Finding(
            kind="columns",
            title=f"{len(strong_unknown)} sütun uygulamada karşılıksız",
            detail=(
                "Bu başlıklar dosyalarınızda tekrar ediyor ama uygulamanın tuttuğu alanlar "
                "arasında yok; bu bilgiler şu an yalnızca Excel'de duruyor."
            ),
            evidence=[listed],
            suggestion=(
                "Yolcu kaydına şu alanları ekle ve hem formda hem listede göster: "
                + ", ".join(name for name, _count in strong_unknown)
            ),
            weight=sum(count for _name, count in strong_unknown),
        ))

    documents = by_kind.get("belge", 0)
    images = by_kind.get("gorsel", 0)
    if documents + images >= 50 and dated_folders >= 2:
        findings.append(Finding(
            kind="archive",
            title=f"{documents} belge ve {images} görsel tarihli klasörlerde duruyor",
            detail=(
                "Dosya düzeniniz zaten güne göre; uygulamanın kayıt klasörleri bunu karşılıyor "
                "ama içe aktarma bu ağaçtan toplu yapılamıyor."
            ),
            evidence=[f"Tarih adlı klasör sayısı: {dated_folders}"],
            suggestion=(
                "Klasör ağacından toplu içe aktarma ekle: seçilen kök klasördeki tarih "
                "klasörlerini gezip PDF ve JPG'leri ilgili günün kaydına bağlasın."
            ),
            weight=documents + images,
        ))

    if by_kind.get("tablo", 0) >= 20 and len(templates) > 5:
        findings.append(Finding(
            kind="consolidate",
            title=f"{len(templates)} farklı tablo düzeni var",
            detail=(
                "Aynı işin birden çok şablonla yürüdüğü anlamına gelir. Uygulamada tek bir "
                "kayıt tipi altında toplanırsa raporlar da tek yerden çıkar."
            ),
            evidence=[
                f"En yaygın {index + 1}. düzen: {group.count} dosya"
                for index, group in enumerate(templates[:3])
            ],
            suggestion=(
                "İçe aktarma sırasında farklı sütun adlarını aynı alana eşleyen bir eşleme "
                "ekranı ekle; eşlemeler hatırlansın ve sonraki dosyalarda otomatik uygulansın."
            ),
            weight=len(templates),
        ))

    findings.sort(key=lambda finding: finding.weight, reverse=True)
    return findings
